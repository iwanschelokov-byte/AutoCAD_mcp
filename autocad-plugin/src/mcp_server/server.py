"""
AutoCAD MCP Server

Exposes AutoCAD operations as MCP tools that AI assistants (Claude) can call.
Communicates with the AutoCAD .NET plugin via TCP socket using JSON-RPC 2.0.

Architecture:
    Claude (MCP Client) -> stdio -> This Server -> TCP socket -> AutoCAD Plugin -> AutoCAD API

Handles:
    Every `handle` in and out of these tools is the hexadecimal entity handle,
    the same string the properties palette, the LIST command, DXF group 5 and
    AutoLISP's (handent "...") use - e.g. "97B176". Handles emitted by older
    builds of this plugin were decimal ("9941366") and are still accepted on
    input, but they are no longer produced.
"""

import os
import json
import asyncio
import logging
from mcp.server.fastmcp import FastMCP
from autocad_client import get_client

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("autocad-mcp")

HOST = os.environ.get("AUTOCAD_MCP_HOST", "localhost")
PORT = int(os.environ.get("AUTOCAD_MCP_PORT", "8081"))

mcp = FastMCP("AutoCAD MCP Server")


# =============================================================================
# Helper
# =============================================================================

async def _raw(method: str, params: dict | None = None):
    """Send a command to AutoCAD and return the decoded result object."""
    client = await get_client(HOST, PORT)
    return await client.send_command(method, params)


async def _call(method: str, params: dict | None = None) -> str:
    """Send a command to AutoCAD and return the result as formatted text."""
    return json.dumps(await _raw(method, params), indent=2)


def _field(result, name):
    """Read a field from a plugin result, whether or not it is wrapped in 'data'."""
    if isinstance(result, dict):
        if name in result:
            return result[name]
        data = result.get("data")
        if isinstance(data, dict):
            return data.get(name)
    return None


# =============================================================================
# System Tools
# =============================================================================

@mcp.tool()
async def system_status() -> str:
    """Get the AutoCAD plugin status, version, and active document info."""
    return await _call("system_status")


@mcp.tool()
async def list_methods() -> str:
    """List all available methods/commands the AutoCAD plugin supports."""
    return await _call("list_methods")


@mcp.tool()
async def set_system_variable(name: str, value: float | int | str) -> str:
    """Set an AutoCAD system variable (e.g., DIMTXT, LTSCALE, OSMODE). Value can be number or string."""
    return await _call("set_system_variable", {"name": name, "value": value})


@mcp.tool()
async def get_system_variable(name: str) -> str:
    """Get the current value of an AutoCAD system variable."""
    return await _call("get_system_variable", {"name": name})


@mcp.tool()
async def execute_command(
    command: str,
    inputs: list[str] | None = None,
    check: bool = True,
    wait: float = 0.5,
) -> str:
    """Execute a raw AutoCAD command, optionally with all of its interactive inputs.

    The command and every item of `inputs` are joined into ONE space-separated
    string and sent to AutoCAD in a single call. Interactive / multi-step
    commands MUST be passed this way (command + all prompt responses together);
    splitting them across several execute_command calls does not work, because
    AutoCAD cancels a command that is still awaiting input when the next string
    is queued.

    The command runs asynchronously, so the send itself cannot report success.
    With check=True (default) this tool waits `wait` seconds and then reads the
    command log back, so the answer contains what AutoCAD actually did — an
    unknown command or a rejected input shows up under "outcome" instead of
    failing silently. Set check=False for fire-and-forget.

    Examples:
        execute_command("ZOOM E")
        execute_command("_.CIRCLE", ["100,100", "40"])
        execute_command("_.POLYGON", ["6", "400,250", "_I", "60"])
    """
    params: dict = {"command": command}
    if inputs:
        params["inputs"] = [str(i) for i in inputs]

    sent = await _raw("execute_command", params)
    if not check:
        return json.dumps(sent, indent=2)

    since = _field(sent, "since") or 0
    if wait > 0:
        await asyncio.sleep(min(wait, 10.0))

    try:
        outcome = await _raw("read_command_line", {"since": since, "limit": 20})
    except Exception as exc:  # the plugin may predate read_command_line
        outcome = {"error": f"could not read command log: {exc}"}

    return json.dumps({"sent": sent, "outcome": outcome}, indent=2)


@mcp.tool()
async def read_command_line(since: int = 0, limit: int = 20) -> str:
    """Read AutoCAD's recent command activity and the last command-line prompt.

    Pass the "since" value returned by execute_command to see exactly what that
    call produced: which commands started, ended, failed or were cancelled, plus
    the last line AutoCAD echoed (where messages like "Unknown command" appear).
    Call with since=0 for the whole recent history.
    """
    return await _call("read_command_line", {"since": since, "limit": limit})


# =============================================================================
# Drawing / Document Tools
# =============================================================================

@mcp.tool()
async def drawing_new(template: str = "") -> str:
    """Create a new drawing, optionally from a template file path."""
    return await _call("drawing_new", {"template": template})


@mcp.tool()
async def drawing_open(path: str, read_only: bool = False) -> str:
    """Open an existing .dwg file and make it the active drawing.

    If the file is already open it is simply activated (AutoCAD cannot open a
    file it already holds a lock on); the answer then says already_open=true.
    """
    return await _call("drawing_open", {"path": path, "read_only": read_only})


@mcp.tool()
async def drawing_save(path: str = "", mode: str = "copy") -> str:
    """Save the current drawing.

    With no path: saves in place (QSAVE).
    With a path, `mode` decides what "save as" means:
      - "copy" (default): writes the file, but the editing session keeps
        pointing at the original drawing.
      - "saveas": switches the editing session to the new file, like the
        SAVEAS command in the UI.
    """
    params = {}
    if path:
        params["path"] = path
        params["mode"] = mode
    return await _call("drawing_save", params)


@mcp.tool()
async def drawing_info() -> str:
    """Get info about the current drawing: name, path, entity count, layers."""
    return await _call("drawing_info")


@mcp.tool()
async def drawing_list() -> str:
    """List every drawing currently open in AutoCAD, marking the active one."""
    return await _call("drawing_list")


@mcp.tool()
async def drawing_close(path: str = "", save: bool = False) -> str:
    """Close a drawing. Closes the active drawing unless `path` names another one.

    save defaults to False: the drawing is closed and unsaved changes are
    DISCARDED. Pass save=True to write the file first. Closing is what releases
    AutoCAD's lock on the .dwg, so do it before another tool needs the file.

    The result carries a `status` saying which of those actually happened -
    "saved", "closed_unchanged" (there was nothing to save), "changes_discarded"
    (there was, and it is gone) or "closed" (the drawing was not the active one,
    so DBMOD could not be read and no claim is made either way).
    """
    params: dict = {"save": save}
    if path:
        params["path"] = path
    return await _call("drawing_close", params)


@mcp.tool()
async def close_all(save: bool = False, keep: str = "") -> str:
    """Close every open drawing. `keep` optionally names one to leave open.

    save defaults to False — unsaved changes in every closed drawing are
    DISCARDED.
    """
    params: dict = {"save": save}
    if keep:
        params["keep"] = keep
    return await _call("close_all", params)


# =============================================================================
# Entity Creation Tools
# =============================================================================

@mcp.tool()
async def create_line(
    start: list[float],
    end: list[float],
    layer: str = "",
    color: int = -1
) -> str:
    """Draw a line from start [x,y] to end [x,y]. Optionally set layer and color (0-255 ACI)."""
    params = {"start": start, "end": end}
    if layer: params["layer"] = layer
    if color >= 0: params["color"] = color
    return await _call("create_line", params)


@mcp.tool()
async def create_circle(
    center: list[float],
    radius: float,
    layer: str = "",
    color: int = -1
) -> str:
    """Draw a circle at center [x,y] with given radius."""
    params = {"center": center, "radius": radius}
    if layer: params["layer"] = layer
    if color >= 0: params["color"] = color
    return await _call("create_circle", params)


@mcp.tool()
async def create_arc(
    center: list[float],
    radius: float,
    start_angle: float,
    end_angle: float,
    degrees: bool = True,
    layer: str = ""
) -> str:
    """Draw an arc. Angles in degrees by default. Set degrees=false for radians."""
    params = {
        "center": center, "radius": radius,
        "start_angle": start_angle, "end_angle": end_angle,
        "degrees": degrees
    }
    if layer: params["layer"] = layer
    return await _call("create_arc", params)


@mcp.tool()
async def create_polyline(
    points: list[list[float]],
    closed: bool = False,
    layer: str = ""
) -> str:
    """Draw a polyline through a list of points [[x1,y1], [x2,y2], ...]. Set closed=true to close it."""
    params = {"points": points, "closed": closed}
    if layer: params["layer"] = layer
    return await _call("create_polyline", params)


@mcp.tool()
async def create_rectangle(
    corner1: list[float],
    corner2: list[float],
    layer: str = ""
) -> str:
    """Draw a rectangle from corner1 [x,y] to corner2 [x,y]."""
    params = {"corner1": corner1, "corner2": corner2}
    if layer: params["layer"] = layer
    return await _call("create_rectangle", params)


@mcp.tool()
async def create_ellipse(
    center: list[float],
    major_radius: float,
    minor_radius: float,
    layer: str = ""
) -> str:
    """Draw an ellipse at center with major and minor radii."""
    params = {"center": center, "major_radius": major_radius, "minor_radius": minor_radius}
    if layer: params["layer"] = layer
    return await _call("create_ellipse", params)


@mcp.tool()
async def create_text(
    text: str,
    position: list[float],
    height: float = 2.5,
    rotation: float = 0,
    layer: str = ""
) -> str:
    """Place single-line text at position [x,y]. Height in drawing units, rotation in degrees."""
    params = {"text": text, "position": position, "height": height, "rotation": rotation}
    if layer: params["layer"] = layer
    return await _call("create_text", params)


@mcp.tool()
async def create_mtext(
    text: str,
    position: list[float],
    height: float = 2.5,
    width: float = 0,
    layer: str = ""
) -> str:
    """Place multi-line text at position [x,y]. Width=0 means auto-width."""
    params = {"text": text, "position": position, "height": height, "width": width}
    if layer: params["layer"] = layer
    return await _call("create_mtext", params)


@mcp.tool()
async def create_hatch(
    boundary: list[list[float]],
    pattern: str = "ANSI31",
    scale: float = 1.0,
    layer: str = ""
) -> str:
    """Create a hatch inside a boundary defined by points. Pattern examples: ANSI31, SOLID, DOTS."""
    params = {"boundary": boundary, "pattern": pattern, "scale": scale}
    if layer: params["layer"] = layer
    return await _call("create_hatch", params)


# =============================================================================
# Advanced Entity Creation Tools
# =============================================================================

@mcp.tool()
async def create_spline(
    points: list[list[float]],
    closed: bool = False,
    layer: str = "",
    color: int = -1
) -> str:
    """Draw a smooth spline curve through points [[x1,y1], ...]. Set closed=true for closed spline."""
    params = {"points": points, "closed": closed}
    if layer: params["layer"] = layer
    if color >= 0: params["color"] = color
    return await _call("create_spline", params)


@mcp.tool()
async def create_table(
    position: list[float],
    rows: int = 3,
    columns: int = 3,
    row_height: float = 500,
    column_width: float = 2000,
    title: str = "",
    data: list[list[str]] | None = None,
    layer: str = ""
) -> str:
    """Create a table at position [x,y] with rows/columns. Data is 2D array of cell strings."""
    params: dict = {"position": position, "rows": rows, "columns": columns,
                    "row_height": row_height, "column_width": column_width}
    if title: params["title"] = title
    if data: params["data"] = data
    if layer: params["layer"] = layer
    return await _call("create_table", params)


@mcp.tool()
async def bulk_create(entities: list[dict]) -> str:
    """Create multiple entities in one call for performance. Each item: {type: 'line'|'circle'|'arc'|'polyline'|'rectangle'|'text'|'mtext'|'ellipse', params: {...}}."""
    return await _call("bulk_create", {"entities": entities})


# =============================================================================
# Entity Query & Modification Tools
# =============================================================================

@mcp.tool()
async def list_entities(
    layer: str = "",
    type: str = "",
    limit: int = 500,
    offset: int = 0,
    detailed: bool = False,
    min_point: list[float] | None = None,
    max_point: list[float] | None = None,
    mode: str = "crossing",
) -> str:
    """List entities in model space, with filtering and paging.

    `type` accepts any spelling of the type: "BlockReference",
    "AcDbBlockReference", "INSERT" or the alias "block" all work, and several
    can be given comma-separated ("Line,Arc,LWPOLYLINE").

    Supply min_point and max_point to list only a region — "crossing" (default)
    includes anything the box touches, "window" only what is fully inside.

    The answer carries `total` and `truncated`, so you can tell "that is all of
    them" from "ask for the next page with offset".
    """
    params: dict = {"limit": limit, "offset": offset, "detailed": detailed}
    if layer: params["layer"] = layer
    if type: params["type"] = type
    if min_point and max_point:
        params["min_point"] = min_point
        params["max_point"] = max_point
        params["mode"] = mode
    return await _call("list_entities", params)


@mcp.tool()
async def get_entity(handle: str) -> str:
    """Get detailed info about a specific entity by its handle.

    The handle is hexadecimal, as AutoCAD shows it ("97B176"), so it can be
    pasted straight into (handent "97B176") or a DXF filter.
    """
    return await _call("get_entity", {"handle": handle})


@mcp.tool()
async def get_entities(handles: list[str], detailed: bool = True) -> str:
    """Get info about several entities at once, by handle. One call instead of N."""
    return await _call("get_entities", {"handles": [str(h) for h in handles], "detailed": detailed})


@mcp.tool()
async def erase_entity(handle: str) -> str:
    """Delete an entity by its handle ID."""
    return await _call("erase_entity", {"handle": handle})


@mcp.tool()
async def move_entity(
    handle: str,
    from_point: list[float],
    to_point: list[float]
) -> str:
    """Move an entity from one point to another."""
    return await _call("move_entity", {"handle": handle, "from": from_point, "to": to_point})


@mcp.tool()
async def copy_entity(
    handle: str,
    from_point: list[float],
    to_point: list[float]
) -> str:
    """Copy an entity from one point to another. Returns the new entity's handle."""
    return await _call("copy_entity", {"handle": handle, "from": from_point, "to": to_point})


@mcp.tool()
async def rotate_entity(
    handle: str,
    base_point: list[float],
    angle: float
) -> str:
    """Rotate an entity around a base point by an angle in degrees."""
    return await _call("rotate_entity", {"handle": handle, "base_point": base_point, "angle": angle})


@mcp.tool()
async def scale_entity(
    handle: str,
    base_point: list[float],
    factor: float
) -> str:
    """Scale an entity from a base point by a scale factor."""
    return await _call("scale_entity", {"handle": handle, "base_point": base_point, "factor": factor})


@mcp.tool()
async def mirror_entity(
    handle: str,
    mirror_line_start: list[float],
    mirror_line_end: list[float],
    erase_source: bool = False
) -> str:
    """Mirror an entity across a line. Set erase_source=true to remove the original."""
    return await _call("mirror_entity", {
        "handle": handle,
        "mirror_line_start": mirror_line_start,
        "mirror_line_end": mirror_line_end,
        "erase_source": erase_source
    })


@mcp.tool()
async def set_entity_properties(
    handle: str,
    layer: str = "",
    color: int = -1,
    linetype: str = "",
    lineweight: int = -1
) -> str:
    """Change properties of an existing entity: layer, color (ACI 0-255), linetype, lineweight."""
    params: dict = {"handle": handle}
    if layer: params["layer"] = layer
    if color >= 0: params["color"] = color
    if linetype: params["linetype"] = linetype
    if lineweight >= 0: params["lineweight"] = lineweight
    return await _call("set_entity_properties", params)


@mcp.tool()
async def offset_entity(
    handle: str,
    distance: float,
    side: str = "both"
) -> str:
    """Offset a curve entity (line/arc/polyline/circle) by distance. Side: 'left', 'right', or 'both'."""
    return await _call("offset_entity", {"handle": handle, "distance": distance, "side": side})


@mcp.tool()
async def explode_entity(handle: str, erase_original: bool = True) -> str:
    """Explode a block/polyline/dimension into primitive entities. Returns new entity handles."""
    return await _call("explode_entity", {"handle": handle, "erase_original": erase_original})


@mcp.tool()
async def array_rectangular(
    handle: str,
    rows: int,
    columns: int,
    row_spacing: float,
    column_spacing: float
) -> str:
    """Create rectangular array of entity. Returns handles of new copies (original stays)."""
    return await _call("array_rectangular", {
        "handle": handle, "rows": rows, "columns": columns,
        "row_spacing": row_spacing, "column_spacing": column_spacing
    })


@mcp.tool()
async def array_polar(
    handle: str,
    center: list[float],
    count: int,
    total_angle: float = 360,
    rotate_items: bool = True
) -> str:
    """Create polar (circular) array of entity around center point."""
    return await _call("array_polar", {
        "handle": handle, "center": center, "count": count,
        "total_angle": total_angle, "rotate_items": rotate_items
    })


@mcp.tool()
async def join_entities(handles: list[str]) -> str:
    """Join contiguous lines/arcs into a single polyline. First handle must be a polyline."""
    return await _call("join_entities", {"handles": handles})


@mcp.tool()
async def bulk_erase(
    handles: list[str] | None = None,
    layer: str = "",
    type: str = ""
) -> str:
    """Erase multiple entities at once. By handles list, or by layer/type filter."""
    params: dict = {}
    if handles: params["handles"] = handles
    if layer: params["layer"] = layer
    if type: params["type"] = type
    return await _call("bulk_erase", params)


@mcp.tool()
async def undo_last(count: int = 1) -> str:
    """Undo the last N operations in AutoCAD."""
    return await _call("undo_last", {"count": count})


# =============================================================================
# Query & Measurement Tools
# =============================================================================

@mcp.tool()
async def measure_distance(point1: list[float], point2: list[float]) -> str:
    """Measure distance between two points. Returns distance, dx, dy, angle."""
    return await _call("measure_distance", {"point1": point1, "point2": point2})


@mcp.tool()
async def measure_area(handle: str) -> str:
    """Measure area of a closed entity (polyline, circle, ellipse, hatch). Returns area and perimeter."""
    return await _call("measure_area", {"handle": handle})


@mcp.tool()
async def get_bounding_box(handle: str) -> str:
    """Get the bounding box of an entity. Returns min_point, max_point, width, height."""
    return await _call("get_bounding_box", {"handle": handle})


@mcp.tool()
async def select_by_window(
    min_point: list[float],
    max_point: list[float],
    limit: int = 500,
    mode: str = "window",
    layer: str = "",
    type: str = "",
    offset: int = 0,
    detailed: bool = False,
) -> str:
    """Find entities in a rectangular region.

    mode="window" (default) returns only entities fully inside the box;
    mode="crossing" also returns anything the box touches — use it when picking
    a title block or a sheet region, whose entities stick out past the frame.

    Optional layer/type filters narrow the result; `type` accepts "INSERT",
    "AcDbBlockReference", "BlockReference" or "block" alike. The answer carries
    `total` and `truncated` so a cut-off list is never mistaken for a complete one.
    """
    params: dict = {
        "min_point": min_point,
        "max_point": max_point,
        "limit": limit,
        "offset": offset,
        "mode": mode,
        "detailed": detailed,
    }
    if layer: params["layer"] = layer
    if type: params["type"] = type
    return await _call("select_by_window", params)


@mcp.tool()
async def select_by_properties(
    layer: str = "",
    type: str = "",
    color: int = -1,
    linetype: str = "",
    limit: int = 500,
    offset: int = 0,
    detailed: bool = False,
    block_name: str = "",
) -> str:
    """Find entities matching property filters (AND logic).

    `type` accepts the .NET name ("BlockReference"), the AutoCAD class name
    ("AcDbBlockReference"), the DXF name ("INSERT") or an alias ("block",
    "text", "anytext", "pline", "dimension", "curve"), and a comma-separated
    list of any of those. `block_name` further restricts block references to
    one block definition.
    """
    params: dict = {"limit": limit, "offset": offset, "detailed": detailed}
    if layer: params["layer"] = layer
    if type: params["type"] = type
    if color >= 0: params["color"] = color
    if linetype: params["linetype"] = linetype
    if block_name: params["block_name"] = block_name
    return await _call("select_by_properties", params)


@mcp.tool()
async def find_intersections(handle1: str, handle2: str) -> str:
    """Find intersection points between two curve entities."""
    return await _call("find_intersections", {"handle1": handle1, "handle2": handle2})


@mcp.tool()
async def search_text(keyword: str, case_sensitive: bool = False, limit: int = 100) -> str:
    """Search ALL text in the drawing (DBText, MText, block attributes, block names) for a keyword. Returns matching text with positions. Use this to find rooms, labels, equipment, annotations by name."""
    return await _call("search_text", {"keyword": keyword, "case_sensitive": case_sensitive, "limit": limit})


@mcp.tool()
async def find_nearest(
    point: list[float],
    radius: float = 0,
    type: str = "",
    layer: str = "",
    limit: int = 20
) -> str:
    """Find entities nearest to a point [x,y], sorted by distance. Filter by type/layer. Set radius=0 for unlimited range."""
    params: dict = {"point": point, "limit": limit}
    if radius > 0:
        params["radius"] = radius
    if type:
        params["type"] = type
    if layer:
        params["layer"] = layer
    return await _call("find_nearest", params)


@mcp.tool()
async def measure_between(handle1: str, handle2: str) -> str:
    """Measure distance between two entities (center-to-center and closest approach). Returns dx, dy, distances, and entity descriptions."""
    return await _call("measure_between", {"handle1": handle1, "handle2": handle2})


# =============================================================================
# Layer Tools
# =============================================================================

@mcp.tool()
async def list_layers() -> str:
    """List all layers with their properties (color, frozen, locked, current)."""
    return await _call("list_layers")


@mcp.tool()
async def create_layer(
    name: str,
    color: int = 7,
    set_current: bool = False,
    linetype: str = ""
) -> str:
    """Create a new layer with optional color (ACI 0-255) and linetype."""
    params = {"name": name, "color": color, "set_current": set_current}
    if linetype: params["linetype"] = linetype
    return await _call("create_layer", params)


@mcp.tool()
async def set_current_layer(name: str) -> str:
    """Set the active/current layer by name."""
    return await _call("set_current_layer", {"name": name})


@mcp.tool()
async def set_layer_properties(
    name: str,
    color: int = -1,
    is_off: bool | None = None,
    is_frozen: bool | None = None,
    is_locked: bool | None = None
) -> str:
    """Modify layer properties. Only specified parameters are changed."""
    params: dict = {"name": name}
    if color >= 0: params["color"] = color
    if is_off is not None: params["is_off"] = is_off
    if is_frozen is not None: params["is_frozen"] = is_frozen
    if is_locked is not None: params["is_locked"] = is_locked
    return await _call("set_layer_properties", params)


@mcp.tool()
async def delete_layer(name: str) -> str:
    """Delete a layer. Entities on it are moved to layer '0'. Cannot delete '0' or current layer."""
    return await _call("delete_layer", {"name": name})


@mcp.tool()
async def rename_layer(old_name: str, new_name: str) -> str:
    """Rename a layer. Cannot rename layer '0'."""
    return await _call("rename_layer", {"old_name": old_name, "new_name": new_name})


# =============================================================================
# Style Tools
# =============================================================================

@mcp.tool()
async def create_dimension_style(
    name: str,
    text_height: float = 0,
    arrow_size: float = 0,
    linear_scale_factor: float = 0,
    decimal_places: int = -1,
    text_above: bool | None = None,
    text_color: int = -1,
    dim_line_color: int = -1,
    ext_line_color: int = -1,
    suffix: str = "",
    text_style: str = "",
    extension_offset: float = 0,
    extension_extend: float = 0,
    text_gap: float = 0,
    set_current: bool = False
) -> str:
    """Create or update a dimension style. Fixes invisible dim text by setting proper text_height and arrow_size."""
    params: dict = {"name": name, "set_current": set_current}
    if text_height > 0: params["text_height"] = text_height
    if arrow_size > 0: params["arrow_size"] = arrow_size
    if linear_scale_factor > 0: params["linear_scale_factor"] = linear_scale_factor
    if decimal_places >= 0: params["decimal_places"] = decimal_places
    if text_above is not None: params["text_above"] = text_above
    if text_color >= 0: params["text_color"] = text_color
    if dim_line_color >= 0: params["dim_line_color"] = dim_line_color
    if ext_line_color >= 0: params["ext_line_color"] = ext_line_color
    if suffix: params["suffix"] = suffix
    if text_style: params["text_style"] = text_style
    if extension_offset > 0: params["extension_offset"] = extension_offset
    if extension_extend > 0: params["extension_extend"] = extension_extend
    if text_gap > 0: params["text_gap"] = text_gap
    return await _call("create_dimension_style", params)


@mcp.tool()
async def create_text_style(
    name: str,
    font: str = "Arial",
    height: float = 0,
    width_factor: float = 1.0,
    oblique_angle: float = 0,
    set_current: bool = False
) -> str:
    """Create or update a text style. Height=0 means variable (uses DIMTXT for dimensions)."""
    return await _call("create_text_style", {
        "name": name, "font": font, "height": height,
        "width_factor": width_factor, "oblique_angle": oblique_angle, "set_current": set_current
    })


@mcp.tool()
async def list_dimension_styles() -> str:
    """List all dimension styles with their properties."""
    return await _call("list_dimension_styles")


@mcp.tool()
async def list_text_styles() -> str:
    """List all text styles with their font, height, and width factor."""
    return await _call("list_text_styles")


# =============================================================================
# Block Tools
# =============================================================================

@mcp.tool()
async def list_blocks() -> str:
    """List all block definitions in the drawing with their attributes."""
    return await _call("list_blocks")


@mcp.tool()
async def create_block(
    name: str,
    handles: list[str],
    base_point: list[float] | None = None,
    erase_originals: bool = False
) -> str:
    """Create a new block definition from existing entities. Provide entity handles and a block name."""
    params: dict = {"name": name, "handles": handles, "erase_originals": erase_originals}
    if base_point: params["base_point"] = base_point
    return await _call("create_block", params)


@mcp.tool()
async def insert_block(
    name: str,
    position: list[float],
    rotation: float = 0,
    scale_x: float = 1.0,
    scale_y: float = 1.0,
    scale_z: float = 1.0,
    layer: str = "",
    attributes: dict | None = None
) -> str:
    """Insert a block by name at position [x,y]. Rotation in degrees. Attributes as {TAG: value}."""
    params: dict = {
        "name": name, "position": position, "rotation": rotation,
        "scale_x": scale_x, "scale_y": scale_y, "scale_z": scale_z
    }
    if layer: params["layer"] = layer
    if attributes: params["attributes"] = attributes
    return await _call("insert_block", params)


@mcp.tool()
async def import_block(
    source_path: str,
    block_names: list[str] | None = None
) -> str:
    """Import block definitions from an external DWG file into the current drawing.
    Use this to bring in blocks from other drawings without needing Design Center.
    After importing, use insert_block to place them. If block_names is omitted, all user blocks are imported."""
    params: dict = {"source_path": source_path}
    if block_names: params["block_names"] = block_names
    return await _call("import_block", params)


# =============================================================================
# Annotation Tools
# =============================================================================

@mcp.tool()
async def create_linear_dimension(
    point1: list[float],
    point2: list[float],
    dimension_line_position: list[float],
    rotation: float = 0,
    text: str = "",
    layer: str = ""
) -> str:
    """Create a linear (horizontal/vertical) dimension between two points."""
    params: dict = {
        "point1": point1, "point2": point2,
        "dimension_line_position": dimension_line_position,
        "rotation": rotation
    }
    if text: params["text"] = text
    if layer: params["layer"] = layer
    return await _call("create_linear_dimension", params)


@mcp.tool()
async def create_angular_dimension(
    center: list[float],
    point1: list[float],
    point2: list[float],
    dimension_arc_position: list[float],
    text: str = "",
    layer: str = ""
) -> str:
    """Create an angular dimension measuring the angle at center between two lines."""
    params: dict = {"center": center, "point1": point1, "point2": point2,
                    "dimension_arc_position": dimension_arc_position}
    if text: params["text"] = text
    if layer: params["layer"] = layer
    return await _call("create_angular_dimension", params)


@mcp.tool()
async def create_radial_dimension(
    center: list[float],
    chord_point: list[float],
    leader_length: float = 0,
    text: str = "",
    layer: str = ""
) -> str:
    """Create a radial dimension for a circle or arc."""
    params: dict = {"center": center, "chord_point": chord_point, "leader_length": leader_length}
    if text: params["text"] = text
    if layer: params["layer"] = layer
    return await _call("create_radial_dimension", params)


@mcp.tool()
async def create_diameter_dimension(
    center: list[float],
    chord_point: list[float],
    leader_length: float = 0,
    text: str = "",
    layer: str = ""
) -> str:
    """Create a diameter dimension for a circle."""
    params: dict = {"center": center, "chord_point": chord_point, "leader_length": leader_length}
    if text: params["text"] = text
    if layer: params["layer"] = layer
    return await _call("create_diameter_dimension", params)


@mcp.tool()
async def create_leader(
    points: list[list[float]],
    text: str = "",
    text_height: float = 2.5,
    layer: str = ""
) -> str:
    """Create a leader (callout arrow) with text. Points: [[arrow_tip], ..., [landing_point]]."""
    params: dict = {"points": points, "text": text, "text_height": text_height}
    if layer: params["layer"] = layer
    return await _call("create_leader", params)


@mcp.tool()
async def create_aligned_dimension(
    point1: list[float],
    point2: list[float],
    dimension_line_position: list[float],
    text: str = "",
    layer: str = ""
) -> str:
    """Create an aligned dimension between two points (measures along the line between them)."""
    params: dict = {
        "point1": point1, "point2": point2,
        "dimension_line_position": dimension_line_position
    }
    if text: params["text"] = text
    if layer: params["layer"] = layer
    return await _call("create_aligned_dimension", params)


# =============================================================================
# View Tools
# =============================================================================

@mcp.tool()
async def zoom_extents() -> str:
    """Zoom to show all entities in the drawing."""
    return await _call("zoom_extents")


@mcp.tool()
async def zoom_window(min_point: list[float], max_point: list[float]) -> str:
    """Zoom to a specific rectangular area defined by min [x,y] and max [x,y] corners."""
    return await _call("zoom_window", {"min": min_point, "max": max_point})


# =============================================================================
# Drawing Utility Tools
# =============================================================================

@mcp.tool()
async def purge_drawing() -> str:
    """Remove all unused layers, blocks, styles, and linetypes from the drawing."""
    return await _call("purge_drawing")


@mcp.tool()
async def set_units(
    linear_units: int = -1,
    precision: int = -1,
    insert_units: int = -1,
    angle_units: int = -1,
    angle_precision: int = -1
) -> str:
    """Set drawing units. linear_units: 1=Scientific,2=Decimal,3=Engineering,4=Architectural. insert_units: 4=mm,5=cm,6=m."""
    params: dict = {}
    if linear_units >= 0: params["linear_units"] = linear_units
    if precision >= 0: params["precision"] = precision
    if insert_units >= 0: params["insert_units"] = insert_units
    if angle_units >= 0: params["angle_units"] = angle_units
    if angle_precision >= 0: params["angle_precision"] = angle_precision
    return await _call("set_units", params)


# =============================================================================
# Plotting
# =============================================================================

_PT_PER_MM = 72.0 / 25.4


def _trim_pdf_to_size(path: str, need_w: float, need_h: float, tol: float = 0.1) -> dict:
    """Crop every page of `path` down to need_w x need_h millimetres, centred.

    AutoCAD can only plot onto a paper size its driver defines, so a
    non-standard sheet (840x594, 297x630) has to be plotted 1:1 centred on the
    nearest larger sheet and the surplus removed afterwards. That is what this
    does: it rewrites the PDF MediaBox to a box of exactly the requested size,
    concentric with the page that came out of the driver, and drops the other
    boxes so no viewer falls back to them.

    The page size is measured from the file rather than taken from the plot
    report, because PDF drivers quantize the sheet: "DWG To PDF.pc3" lands on a
    0.0423 mm grid, so a nominal 841x594 A1 arrives as 841.02 x 594.08.

    Returns a dict that is merged into the plot result. It always contains
    `trimmed`; on failure it also contains `trim_error` with the reason, and
    never raises - a PDF that could not be trimmed is still a valid plot.
    """
    try:
        import pikepdf
    except ImportError:
        import sys
        return {
            "trimmed": False,
            "trim_error": (
                "pikepdf is not installed for the Python running this MCP "
                "server (" + sys.executable + "), so the page was left at the "
                "printer sheet size. Install it with: \"" + sys.executable +
                "\" -m pip install pikepdf   - then restart the MCP host. The "
                "plot itself is correct and centred on the sheet, so meanwhile "
                "it can be cropped by hand to the `trim_target_mm` below."
            ),
            "trim_target_mm": [round(need_w, 2), round(need_h, 2)],
        }

    tmp = path + ".trim.tmp"
    try:
        with pikepdf.open(path) as pdf:
            pages = 0
            box_mm = None
            size_mm = None
            for page in pdf.pages:
                mb = [float(v) for v in page.mediabox]
                x0, y0 = min(mb[0], mb[2]), min(mb[1], mb[3])
                x1, y1 = max(mb[0], mb[2]), max(mb[1], mb[3])
                pw, ph = (x1 - x0) / _PT_PER_MM, (y1 - y0) / _PT_PER_MM

                # The driver may have rotated the sheet, so try both ways round
                # and keep whichever actually fits inside the page.
                fits = [(w, h) for w, h in ((need_w, need_h), (need_h, need_w))
                        if w <= pw + tol and h <= ph + tol]
                if not fits:
                    return {
                        "trimmed": False,
                        "trim_error": (
                            "the target %.2f x %.2f mm does not fit inside the "
                            "%.2f x %.2f mm page, so nothing was cropped - the "
                            "plot may have been scaled or rotated unexpectedly"
                            % (need_w, need_h, pw, ph)
                        ),
                        "trim_target_mm": [round(need_w, 2), round(need_h, 2)],
                    }
                tw, th = min(fits, key=lambda wh: (pw - wh[0]) + (ph - wh[1]))

                if (pw - tw) <= tol and (ph - th) <= tol:
                    return {
                        "trimmed": False,
                        "trim_reason": (
                            "the page is already %.2f x %.2f mm, within %.2f mm "
                            "of the target - nothing to crop" % (pw, ph, tol)
                        ),
                        "trimmed_size_mm": [round(pw, 2), round(ph, 2)],
                    }

                nx0 = x0 + (x1 - x0 - tw * _PT_PER_MM) / 2.0
                ny0 = y0 + (y1 - y0 - th * _PT_PER_MM) / 2.0
                box = [nx0, ny0, nx0 + tw * _PT_PER_MM, ny0 + th * _PT_PER_MM]
                page.mediabox = box
                # A leftover CropBox at the old size would win in most viewers,
                # and the other three would misreport the trimmed sheet.
                for key in ("/CropBox", "/TrimBox", "/ArtBox", "/BleedBox"):
                    if key in page:
                        del page[key]
                pages += 1
                if box_mm is None:
                    box_mm = [round(v / _PT_PER_MM, 2) for v in box]
                    size_mm = [round(tw, 2), round(th, 2)]

            if not pages:
                return {"trimmed": False, "trim_error": "the PDF has no pages"}
            pdf.save(tmp)
        os.replace(tmp, path)
        return {
            "trimmed": True,
            "trimmed_size_mm": size_mm,
            "trim_box_mm": box_mm,
            "trimmed_pages": pages,
        }
    except Exception as e:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except OSError:
            pass
        return {
            "trimmed": False,
            "trim_error": "%s: %s - the plot itself is intact, only the crop "
                          "failed" % (type(e).__name__, e),
            "trim_target_mm": [round(need_w, 2), round(need_h, 2)],
        }


def _merge(result, extra: dict):
    """Merge extra fields into a plugin result, next to the fields it returned."""
    if isinstance(result, dict):
        data = result.get("data")
        (data if isinstance(data, dict) else result).update(extra)
    return result


def _restate_size(result, path: str) -> None:
    """Replace a stale `file_size` (and the byte count inside `message`) with
    the size the file has now. Called after the PDF is rewritten in place; a
    file that cannot be stat-ed is left reporting what the plugin said."""
    old_size = _field(result, "file_size")
    try:
        new_size = os.path.getsize(path)
    except OSError:
        return
    if not isinstance(old_size, int) or new_size == old_size:
        return
    patch = {"file_size": new_size}
    message = _field(result, "message")
    if isinstance(message, str):
        stale = f"{old_size:,}"
        if stale in message:
            patch["message"] = message.replace(stale, f"{new_size:,}")
    _merge(result, patch)


@mcp.tool()
async def plot_devices(device: str = "", plotter: str = "", filter: str = "") -> str:
    """List plot devices, plot style tables, and a device's paper sizes in mm.

    Canonical paper names are not guessable - the A1 sheet a localized AutoCAD
    shows as "ISO A1 (841.00 x 594.00 mm)" is called
    "ISO_full_bleed_A1_(841.00_x_594.00_MM)" in the API. Pass `device` (e.g.
    "DWG To PDF.pc3") to get its media list with sheet size, printable area and
    margins in millimetres, so plot_to_pdf(paper=...) can be given a name that
    exists. `filter` narrows the media list by substring, e.g. "A1".

    `plotter` is an accepted synonym for `device`. Some MCP hosts route calls
    through a bridge that reserves the argument name "device" for itself and
    consumes it before the tool ever sees it; on those hosts `device` silently
    arrives empty and `plotter` is the way through. Pass one or the other -
    `device` wins if both are given.

    Called with no device, it lists the sheets of "DWG To PDF.pc3". The answer
    always echoes `requested_device` and `requested_filter`, so a dropped
    argument is visible rather than silently ignored.
    """
    dev = device or plotter
    params: dict = {}
    if dev: params["device"] = dev
    # Forwarded as well as merged, so the plugin can echo it back honestly.
    if plotter: params["plotter"] = plotter
    if filter: params["filter"] = filter
    return await _call("plot_devices", params)


@mcp.tool()
async def plot_to_pdf(
    output_path: str,
    device: str = "",
    plotter: str = "",
    paper: str = "",
    style_table: str = "",
    area: str = "",
    window: list[float] = None,
    scale: str = "",
    offset: str = "",
    orientation: str = "",
    layout: str = "",
    lineweights: bool = True,
    overwrite: bool = True,
    trim: bool = True,
) -> str:
    """Plot the drawing to PDF and wait for the file to be written.

    Goes through the PlottingServices API rather than the command line, so it
    reports the real outcome and the real file size instead of fire-and-forget.

    output_path  Where to write. A relative path resolves next to the drawing.
    device       Plot device, default "DWG To PDF.pc3". See plot_devices.
    plotter      Synonym for `device`, for MCP hosts whose bridge reserves the
                 argument name "device" and eats it before the tool sees it.
                 `device` wins if both are given.
    paper        Canonical or localized media name, or "auto" (default): the
                 smallest sheet whose *printable* area fits the window at the
                 requested scale, which is what picks full-bleed sheets over
                 bordered ones when the frame reaches the paper edge.
    style_table  Plot style table, default "monochrome.ctb"; "none" keeps
                 whatever the layout already uses.
    area         extents | window | display | layout | limits. Defaults to
                 "window" when `window` is given, otherwise "extents".
    window       [x1, y1, x2, y2] in drawing units - normally the sheet frame,
                 which in these drawings is a closed LWPOLYLINE on Defpoints.
    scale        "1=1" (default), "1:100", a number, or "fit".
    offset       "center" (default) or "dx,dy" in millimetres.
    orientation  "auto" (default), "portrait" or "landscape".
    layout       Layout to plot; default the current one.
    lineweights  Honour lineweights, default True.
    overwrite    Replace an existing file, default True.
    trim         Crop the finished page down to the plotted window, default
                 True. Set False to keep the whole printer sheet.

    The result reports the media actually used and its millimetre size.

    AutoCAD can only plot onto a paper size the device defines, so a
    non-standard sheet (840x594, 297x630) is plotted 1:1 centred on the nearest
    larger sheet and the surplus is cropped afterwards, here in the MCP server,
    using pikepdf. That crop needs `pikepdf` installed (it is in
    requirements.txt) and only happens for a `window` plot at a fixed scale with
    `offset` left at "center".

    The answer always says what happened: `trimmed` true or false, with
    `trimmed_size_mm` and `trim_box_mm` when it worked, and `trim_error` or
    `trim_reason` with the text when it did not - a missing pikepdf, a page that
    is already the right size, or a target that does not fit. It is never
    swallowed, and a failed crop never invalidates the plot.
    """
    params: dict = {"output_path": output_path, "lineweights": lineweights,
                    "overwrite": overwrite}
    dev = device or plotter
    if dev: params["device"] = dev
    if plotter: params["plotter"] = plotter
    if paper: params["paper"] = paper
    if style_table: params["style_table"] = style_table
    if area: params["area"] = area
    if window: params["window"] = window
    if scale: params["scale"] = scale
    if layout: params["layout"] = layout
    if orientation: params["orientation"] = orientation
    if offset:
        s = offset.strip()
        if s.lower() in ("center", "centre"):
            params["offset"] = "center"
        else:
            try:
                dx, dy = (float(p) for p in s.replace(";", ",").split(",")[:2])
            except ValueError:
                return json.dumps({
                    "success": False,
                    "error": f"offset must be 'center' or 'dx,dy' in millimetres, got {offset!r}",
                }, indent=2)
            params["offset"] = [dx, dy]

    result = await _raw("plot_to_pdf", params)

    # Crop the surplus sheet away. Everything needed is in the plot report:
    # `required_mm` is the window multiplied by the scale, i.e. the sheet the
    # drawing actually asked for, and the plugin only fills it in for a window
    # plot at a fixed scale.
    if isinstance(result, dict) and _field(result, "success") is not False:
        need = _field(result, "required_mm")
        used = _field(result, "output_path") or output_path
        where = _field(result, "offset")
        if not trim:
            _merge(result, {"trimmed": False,
                            "trim_reason": "trim=false, the full printer sheet was kept"})
        elif not (isinstance(need, (list, tuple)) and len(need) == 2):
            _merge(result, {"trimmed": False,
                            "trim_reason": "no target size to crop to - cropping "
                                           "applies to a window plot at a fixed scale"})
        elif where != "center":
            _merge(result, {"trimmed": False,
                            "trim_reason": "the plot is offset rather than centred, "
                                           "so the surplus is not symmetric and was left alone"})
        else:
            trim_report = await asyncio.to_thread(
                _trim_pdf_to_size, used, float(need[0]), float(need[1]))
            _merge(result, trim_report)
            if trim_report.get("trimmed"):
                # The plugin measured the file before pikepdf rewrote it, so the
                # size it reported is the size of the untrimmed sheet. Re-stat
                # and correct both the field and the sentence built from it,
                # otherwise the answer names a size the file no longer has.
                _restate_size(result, used)

    return json.dumps(result, indent=2)


# =============================================================================
# Screenshot
# =============================================================================

@mcp.tool()
async def capture_screenshot(output_path: str = None) -> str:
    """Capture a screenshot of the current AutoCAD window and save as PNG.
    Returns the file path of the saved image. Use this to visually verify
    that drawing operations were completed correctly. The AI can view the
    returned image file to check the result."""
    params: dict = {}
    if output_path:
        params["output_path"] = output_path
    return await _call("capture_screenshot", params)


# =============================================================================
# Excel Table Import
# =============================================================================

@mcp.tool()
async def create_table_from_excel(
    excel_path: str,
    position: list[float],
    sheet_name: str = "",
    title: str = "",
    text_height: float = 120,
    header_text_height: float = 140,
    title_text_height: float = 250,
    row_height: float = 350,
    header_row_height: float = 400,
    title_row_height: float = 600,
    min_col_width: float = 2000,
    char_width: float = 80,
    color: int = 3,
    layer: str = "TABLE",
    start_row: int = 1,
    end_row: int = 0,
    start_col: int = 1,
    end_col: int = 0
) -> str:
    """Create a table in AutoCAD from an Excel file. Reads the spreadsheet, auto-sizes columns, and draws grid lines + text.

    Args:
        excel_path: Absolute path to the .xlsx file.
        position: Insertion point [x, y] for top-left corner of the table.
        sheet_name: Sheet name to read (default: first sheet).
        title: Optional title row text spanning full width.
        text_height: Text height for data cells.
        header_text_height: Text height for the first row (header).
        title_text_height: Text height for the title row.
        row_height: Row height for data cells.
        header_row_height: Row height for the first (header) row.
        title_row_height: Row height for the title row.
        min_col_width: Minimum column width.
        char_width: Approximate width per character for auto-sizing.
        color: ACI color index (default 3 = green).
        layer: Layer name for all entities.
        start_row: First Excel row to read (1-based, default 1).
        end_row: Last Excel row to read (0 = last row).
        start_col: First Excel column to read (1-based, default 1).
        end_col: Last Excel column to read (0 = last column).
    """
    import json as _json

    try:
        import openpyxl
    except ImportError:
        return _json.dumps({"error": "openpyxl is not installed. Run: pip install openpyxl"})

    if not os.path.exists(excel_path):
        return _json.dumps({"error": f"File not found: {excel_path}"})

    # --- Read Excel ---
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return _json.dumps({"error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"})
        ws = wb[sheet_name]
    else:
        ws = wb.active

    sr = start_row
    er = end_row if end_row > 0 else ws.max_row
    sc = start_col
    ec = end_col if end_col > 0 else ws.max_column

    rows_data: list[list[str]] = []
    for row in ws.iter_rows(min_row=sr, max_row=er, min_col=sc, max_col=ec, values_only=True):
        rows_data.append([str(v) if v is not None else "" for v in row])

    if not rows_data:
        return _json.dumps({"error": "No data found in the specified range."})

    num_cols = len(rows_data[0])
    num_rows = len(rows_data)

    # --- Auto-size columns ---
    # Use header text height for header row width calc, data text height for data rows
    col_widths: list[float] = []
    for c in range(num_cols):
        max_w = 0.0
        for r in range(num_rows):
            if c < len(rows_data[r]) and rows_data[r][c]:
                txt_len = len(rows_data[r][c])
                # Scale char_width by text height ratio (larger text needs wider columns)
                h = header_text_height if r == 0 else text_height
                w = txt_len * char_width * (h / 120.0) + 400
                max_w = max(max_w, w)
        col_widths.append(max(min_col_width, max_w))

    total_w = sum(col_widths)
    x0, y0 = position[0], position[1]

    # Column X positions
    col_x = [x0]
    for w in col_widths:
        col_x.append(col_x[-1] + w)

    # Total height
    has_title = bool(title)
    total_h = (title_row_height if has_title else 0) + header_row_height + row_height * (num_rows - 1)

    # --- Build entities ---
    entities: list[dict] = []

    # Horizontal lines
    y = y0
    entities.append({"type": "line", "params": {"start": [x0, y], "end": [x0 + total_w, y], "layer": layer, "color": color}})

    if has_title:
        y -= title_row_height
        entities.append({"type": "line", "params": {"start": [x0, y], "end": [x0 + total_w, y], "layer": layer, "color": color}})

    # Header bottom
    y -= header_row_height
    entities.append({"type": "line", "params": {"start": [x0, y], "end": [x0 + total_w, y], "layer": layer, "color": color}})

    # Data rows
    for i in range(1, num_rows):
        y -= row_height
        entities.append({"type": "line", "params": {"start": [x0, y], "end": [x0 + total_w, y], "layer": layer, "color": color}})

    # Vertical lines
    vert_top = y0 - (title_row_height if has_title else 0)
    for idx, cx in enumerate(col_x):
        # Outer edges span full height including title row
        if has_title and (idx == 0 or idx == len(col_x) - 1):
            entities.append({"type": "line", "params": {"start": [cx, y0], "end": [cx, y0 - total_h], "layer": layer, "color": color}})
        else:
            entities.append({"type": "line", "params": {"start": [cx, vert_top], "end": [cx, y0 - total_h], "layer": layer, "color": color}})

    # Title text (centered via middle-center justification)
    if has_title:
        entities.append({"type": "text", "params": {
            "position": [x0 + total_w / 2, y0 - title_row_height / 2],
            "text": title, "height": title_text_height,
            "layer": layer, "color": color, "justification": "middle-center"
        }})

    # Padding from left cell edge
    cell_pad = 100

    # Header row (first data row)
    header_y_start = y0 - (title_row_height if has_title else 0)
    hy = header_y_start - header_row_height * 0.65
    for c in range(num_cols):
        txt = rows_data[0][c] if c < len(rows_data[0]) else ""
        if txt:
            cx = col_x[c] + cell_pad
            entities.append({"type": "text", "params": {
                "position": [cx, hy], "text": txt, "height": header_text_height,
                "layer": layer, "color": color
            }})

    # Data rows
    data_y_start = header_y_start - header_row_height
    for r in range(1, num_rows):
        ry = data_y_start - (r - 1) * row_height - row_height * 0.65
        for c in range(num_cols):
            txt = rows_data[r][c] if c < len(rows_data[r]) else ""
            if txt:
                cx = col_x[c] + cell_pad
                entities.append({"type": "text", "params": {
                    "position": [cx, ry], "text": txt, "height": text_height,
                    "layer": layer, "color": color
                }})

    # --- Send to AutoCAD in batches ---
    batch_size = 500
    all_handles: list[str] = []
    total_created = 0

    for i in range(0, len(entities), batch_size):
        batch = entities[i:i + batch_size]
        result_str = await _call("bulk_create", {"entities": batch})
        result = _json.loads(result_str)
        if "handles" in result:
            all_handles.extend(result["handles"])
            total_created += result.get("count", len(result["handles"]))

    summary = {
        "success": True,
        "message": f"Table created from '{os.path.basename(excel_path)}'",
        "sheet": ws.title,
        "data_rows": num_rows,
        "columns": num_cols,
        "entities_created": total_created,
        "table_width": total_w,
        "table_height": total_h,
        "position": [x0, y0],
    }
    return _json.dumps(summary, indent=2)


# =============================================================================
# Text Measurement (ground-truth bounding boxes from AutoCAD itself)
# =============================================================================

@mcp.tool()
async def measure_text(
    text: str,
    height: float,
    style: str = "",
    width_factor: float = 0.0,
    oblique: float = 0.0,
) -> str:
    """Measure the exact bounding box AutoCAD would give a single-line text string.

    Returns width/height in drawing units. Use this before placing text to lay
    out tables, columns or labels without overlap, instead of estimating.
    """
    params: dict = {"text": text, "height": height}
    if style:
        params["style"] = style
    if width_factor > 0:
        params["width_factor"] = width_factor
    if oblique:
        params["oblique"] = oblique
    return await _call("measure_text", params)


@mcp.tool()
async def measure_texts(items: list[dict]) -> str:
    """Measure many text strings in one round-trip.

    Each item is {"text": str, "height": float, optional "style"/"width_factor"}.
    Far faster than calling measure_text repeatedly when laying out a table.
    """
    return await _call("measure_texts", {"items": items})


# =============================================================================
# Server Introspection & Safety
# =============================================================================

@mcp.tool()
async def get_capabilities() -> str:
    """Report what this plugin build supports: tool count, AutoCAD version range,
    .NET target, which tools are destructive, and the current safety settings.

    Answers instantly even when AutoCAD is busy or showing a modal dialog.
    """
    return await _call("get_capabilities")


@mcp.tool()
async def get_server_options() -> str:
    """Read the current safety posture (read-only mode, destructive confirmation, audit log)."""
    return await _call("get_server_options")


@mcp.tool()
async def set_server_options(
    read_only: bool | None = None,
    confirm_destructive: bool | None = None,
    audit_log: bool | None = None,
) -> str:
    """Change and persist the safety posture.

    read_only=True refuses every drawing-modifying tool (safe inspection mode).
    confirm_destructive=True (default) makes erase/delete/purge tools require
    __confirm=true. audit_log toggles the JSONL activity log.
    """
    params: dict = {}
    if read_only is not None:
        params["read_only"] = read_only
    if confirm_destructive is not None:
        params["confirm_destructive"] = confirm_destructive
    if audit_log is not None:
        params["audit_log"] = audit_log
    return await _call("set_server_options", params)


# =============================================================================
# Layouts & Paper Space
# =============================================================================

@mcp.tool()
async def list_layouts(include_model: bool = False) -> str:
    """List the drawing's layouts (sheets) with paper size, plot device and scale info."""
    return await _call("list_layouts", {"include_model": include_model})


@mcp.tool()
async def create_layout(name: str, set_current: bool = False) -> str:
    """Create a new layout (sheet) tab."""
    return await _call("create_layout", {"name": name, "set_current": set_current})


@mcp.tool()
async def delete_layout(name: str, __confirm: bool = False) -> str:
    """Delete a layout. Destructive — pass __confirm=true to proceed."""
    return await _call("delete_layout", {"name": name, "__confirm": __confirm})


@mcp.tool()
async def rename_layout(name: str, new_name: str) -> str:
    """Rename a layout."""
    return await _call("rename_layout", {"name": name, "new_name": new_name})


@mcp.tool()
async def set_current_layout(name: str) -> str:
    """Switch the active layout/sheet (use 'Model' for model space)."""
    return await _call("set_current_layout", {"name": name})


@mcp.tool()
async def copy_layout(name: str, new_name: str) -> str:
    """Duplicate a layout, including its page setup and viewports."""
    return await _call("copy_layout", {"name": name, "new_name": new_name})


@mcp.tool()
async def get_page_setup(layout: str = "") -> str:
    """Read a layout's page setup: device, paper size, plot type, scale, style table."""
    params: dict = {}
    if layout:
        params["layout"] = layout
    return await _call("get_page_setup", params)


@mcp.tool()
async def set_page_setup(
    layout: str = "",
    device: str = "",
    paper_size: str = "",
    plot_type: str = "",
    scale: str = "",
    scale_to_fit: bool = False,
    centered: bool | None = None,
    rotation: str = "",
    plot_style_table: str = "",
) -> str:
    """Configure a layout's page setup before plotting.

    device: plotter/printer name (see list_plot_devices).
    paper_size: canonical media name (see list_paper_sizes).
    plot_type: Display, Extents, Limits, View, Window or Layout.
    scale: "1:100" style ratio; or set scale_to_fit=true.
    rotation: 0, 90, 180 or 270.
    """
    params: dict = {}
    for k, v in (
        ("layout", layout), ("device", device), ("paper_size", paper_size),
        ("plot_type", plot_type), ("scale", scale), ("rotation", rotation),
        ("plot_style_table", plot_style_table),
    ):
        if v:
            params[k] = v
    if scale_to_fit:
        params["scale_to_fit"] = True
    if centered is not None:
        params["centered"] = centered
    return await _call("set_page_setup", params)


@mcp.tool()
async def list_plot_devices() -> str:
    """List available plotters/printers configured in AutoCAD.

    Superseded by `plot_devices`, which also returns plot style tables and each
    device's paper sizes with printable areas and margins. Prefer that one.
    """
    return await _call("list_plot_devices")


@mcp.tool()
async def list_paper_sizes(layout: str = "", device: str = "") -> str:
    """List paper sizes available for a plot device (canonical + display names).

    Scoped to a layout. For a device-oriented listing that also reports printable
    areas and margins, use `plot_devices`.
    """
    params: dict = {}
    if layout:
        params["layout"] = layout
    if device:
        params["device"] = device
    return await _call("list_paper_sizes", params)


@mcp.tool()
async def list_viewports(layout: str = "") -> str:
    """List paper-space viewports on a layout with their size, scale and lock state."""
    params: dict = {}
    if layout:
        params["layout"] = layout
    return await _call("list_viewports", params)


@mcp.tool()
async def create_viewport(
    center: list[float],
    width: float,
    height: float,
    layout: str = "",
    scale: str = "",
    view_center: list[float] | None = None,
    locked: bool = False,
    layer: str = "",
) -> str:
    """Create a paper-space viewport on a layout.

    center/width/height are in paper space units. scale like "1:100" sets the
    model-to-paper ratio; view_center pans the model view inside the viewport.
    """
    params: dict = {"center": center, "width": width, "height": height}
    if layout:
        params["layout"] = layout
    if scale:
        params["scale"] = scale
    if view_center:
        params["view_center"] = view_center
    if locked:
        params["locked"] = True
    if layer:
        params["layer"] = layer
    return await _call("create_viewport", params)


@mcp.tool()
async def set_viewport_scale(id: str, scale: str) -> str:
    """Set a viewport's scale, e.g. "1:100". Temporarily unlocks a locked viewport."""
    return await _call("set_viewport_scale", {"id": id, "scale": scale})


@mcp.tool()
async def lock_viewport(id: str, locked: bool = True) -> str:
    """Lock or unlock a viewport so its scale cannot be changed by zooming."""
    return await _call("lock_viewport", {"id": id, "locked": locked})


@mcp.tool()
async def plot_layout(output_path: str, layout: str = "") -> str:
    """Set a layout current and queue an EXPORTPDF for it.

    Prefer `plot_to_pdf`: it waits for the PDF to actually be written and can trim
    the page to a plotted window. This tool returns as soon as the command is
    queued, so the file may not exist yet when it replies.
    """
    params: dict = {"output_path": output_path}
    if layout:
        params["layout"] = layout
    return await _call("plot_layout", params)


# =============================================================================
# External References (Xrefs)
# =============================================================================

@mcp.tool()
async def attach_xref(
    path: str,
    name: str = "",
    position: list[float] | None = None,
    scale: float = 1.0,
    rotation: float = 0.0,
    overlay: bool = False,
    layer: str = "",
) -> str:
    """Attach an external DWG as an xref and place a reference to it.

    overlay=True attaches as an overlay (not nested into drawings that xref this one).
    """
    params: dict = {"path": path, "scale": scale, "rotation": rotation, "overlay": overlay}
    if name:
        params["name"] = name
    if position:
        params["position"] = position
    if layer:
        params["layer"] = layer
    return await _call("attach_xref", params)


@mcp.tool()
async def list_xrefs() -> str:
    """List attached xrefs with path, load status, overlay flag and reference count."""
    return await _call("list_xrefs")


@mcp.tool()
async def reload_xref(name: str) -> str:
    """Reload an xref to pick up changes made to the source DWG."""
    return await _call("reload_xref", {"name": name})


@mcp.tool()
async def unload_xref(name: str) -> str:
    """Unload an xref (keeps the attachment, hides the geometry)."""
    return await _call("unload_xref", {"name": name})


@mcp.tool()
async def detach_xref(name: str, __confirm: bool = False) -> str:
    """Detach an xref completely. Destructive — pass __confirm=true to proceed."""
    return await _call("detach_xref", {"name": name, "__confirm": __confirm})


@mcp.tool()
async def bind_xref(name: str, insert_bind: bool = False) -> str:
    """Bind an xref into the drawing, making its geometry permanent.

    insert_bind=True merges names like INSERT; False keeps them prefixed.
    """
    return await _call("bind_xref", {"name": name, "insert_bind": insert_bind})


@mcp.tool()
async def set_xref_path(name: str, path: str, reload: bool = True) -> str:
    """Repoint a broken or moved xref at a new file path."""
    return await _call("set_xref_path", {"name": name, "path": path, "reload": reload})


@mcp.tool()
async def read_external_dwg(
    path: str,
    include_layers: bool = True,
    include_blocks: bool = True,
    include_layouts: bool = True,
    include_entity_counts: bool = True,
) -> str:
    """Inspect a DWG file WITHOUT opening it in AutoCAD.

    Reads it as a side database and reports layers, blocks, xrefs, layouts and
    entity counts. Use this to check a drawing before opening it.
    """
    return await _call("read_external_dwg", {
        "path": path,
        "include_layers": include_layers,
        "include_blocks": include_blocks,
        "include_layouts": include_layouts,
        "include_entity_counts": include_entity_counts,
    })


@mcp.tool()
async def batch_query_dwgs(
    folder: str,
    recursive: bool = False,
    limit: int = 200,
    include_layers: bool = False,
    include_blocks: bool = False,
    include_layouts: bool = True,
    include_entity_counts: bool = False,
) -> str:
    """Scan a folder of DWG files without opening any of them.

    Ideal for audits: "which drawings use layer X", "how many sheets per file".
    Reports which files failed to read rather than aborting the sweep.
    """
    return await _call("batch_query_dwgs", {
        "folder": folder,
        "recursive": recursive,
        "limit": limit,
        "include_layers": include_layers,
        "include_blocks": include_blocks,
        "include_layouts": include_layouts,
        "include_entity_counts": include_entity_counts,
    })


# =============================================================================
# Block Attributes & Dynamic Blocks
# =============================================================================

@mcp.tool()
async def list_block_attributes(name: str) -> str:
    """List the attribute definitions declared by a block (tags, prompts, defaults)."""
    return await _call("list_block_attributes", {"name": name})


@mcp.tool()
async def get_attribute_values(id: str = "", name: str = "") -> str:
    """Read attribute values.

    Pass id for one block reference, or name to read every reference of that
    block (title blocks, door tags, equipment schedules).
    """
    params: dict = {}
    if id:
        params["id"] = id
    if name:
        params["name"] = name
    return await _call("get_attribute_values", params)


@mcp.tool()
async def set_attribute_values(id: str, attributes: dict) -> str:
    """Set attribute values on one block reference, as {"TAG": "value"}.

    Reports which tags were updated and which were not found on the block.
    """
    return await _call("set_attribute_values", {"id": id, "attributes": attributes})


@mcp.tool()
async def sync_attributes(name: str) -> str:
    """Add attributes to existing block references that were added to the block
    definition after they were inserted (equivalent to ATTSYNC)."""
    return await _call("sync_attributes", {"name": name})


@mcp.tool()
async def get_dynamic_block_properties(id: str) -> str:
    """List a dynamic block reference's parameters, current values and allowed values."""
    return await _call("get_dynamic_block_properties", {"id": id})


@mcp.tool()
async def set_dynamic_block_property(id: str, property: str, value: str) -> str:
    """Set one dynamic block parameter (e.g. a visibility state, length or angle)."""
    return await _call("set_dynamic_block_property",
                       {"id": id, "property": property, "value": value})


@mcp.tool()
async def rename_block(name: str, new_name: str) -> str:
    """Rename a block definition."""
    return await _call("rename_block", {"name": name, "new_name": new_name})


@mcp.tool()
async def delete_block_definition(name: str, __confirm: bool = False) -> str:
    """Delete an unused block definition. Refuses if references still exist.
    Destructive — pass __confirm=true to proceed."""
    return await _call("delete_block_definition", {"name": name, "__confirm": __confirm})


@mcp.tool()
async def count_block_references(name: str = "") -> str:
    """Count how many times each block is inserted — a quick bill of materials.
    Pass name to count one block only."""
    params: dict = {}
    if name:
        params["name"] = name
    return await _call("count_block_references", params)


@mcp.tool()
async def export_block_to_file(name: str, path: str) -> str:
    """Export a block definition to its own DWG file (WBLOCK)."""
    return await _call("export_block_to_file", {"name": name, "path": path})


# =============================================================================
# Modify Operations
# =============================================================================

@mcp.tool()
async def break_entity(id: str, points: list[list[float]]) -> str:
    """Split a curve at one or more points, replacing it with the resulting pieces.

    Points are snapped onto the curve, so approximate coordinates are fine.
    """
    return await _call("break_entity", {"id": id, "points": points})


@mcp.tool()
async def reverse_polyline(id: str) -> str:
    """Reverse a curve's direction (start and end swap)."""
    return await _call("reverse_polyline", {"id": id})


@mcp.tool()
async def polyline_edit(
    id: str,
    closed: bool | None = None,
    width: float | None = None,
    elevation: float | None = None,
    add_vertex: list[float] | None = None,
    index: int | None = None,
    remove_vertex: int | None = None,
) -> str:
    """Edit a polyline: open/close it, set constant width or elevation, or add
    and remove vertices."""
    params: dict = {"id": id}
    if closed is not None:
        params["closed"] = closed
    if width is not None:
        params["width"] = width
    if elevation is not None:
        params["elevation"] = elevation
    if add_vertex is not None:
        params["add_vertex"] = add_vertex
    if index is not None:
        params["index"] = index
    if remove_vertex is not None:
        params["remove_vertex"] = remove_vertex
    return await _call("polyline_edit", params)


@mcp.tool()
async def set_draworder(ids: list[str], position: str = "top", reference_id: str = "") -> str:
    """Change draw order so entities render in front of or behind others.

    position: top, bottom, above or below. above/below also need reference_id.
    """
    params: dict = {"ids": ids, "position": position}
    if reference_id:
        params["reference_id"] = reference_id
    return await _call("set_draworder", params)


@mcp.tool()
async def flatten_entities(ids: list[str] | None = None, z: float = 0.0) -> str:
    """Flatten entities to a single Z elevation. Omit ids to flatten all of model space."""
    params: dict = {"z": z}
    if ids:
        params["ids"] = ids
    return await _call("flatten_entities", params)


@mcp.tool()
async def divide_entity(id: str, segments: int, block: str = "", layer: str = "") -> str:
    """Place markers dividing a curve into N equal segments (AutoCAD DIVIDE).

    Places points by default, or copies of a block if block is given.
    """
    params: dict = {"id": id, "segments": segments}
    if block:
        params["block"] = block
    if layer:
        params["layer"] = layer
    return await _call("divide_entity", params)


@mcp.tool()
async def measure_entity(id: str, interval: float, block: str = "", layer: str = "") -> str:
    """Place markers at a fixed interval along a curve (AutoCAD MEASURE)."""
    params: dict = {"id": id, "interval": interval}
    if block:
        params["block"] = block
    if layer:
        params["layer"] = layer
    return await _call("measure_entity", params)


@mcp.tool()
async def create_region(ids: list[str], erase_source: bool = False, layer: str = "") -> str:
    """Build region(s) from closed loops of curves — the input for extrude/revolve."""
    params: dict = {"ids": ids, "erase_source": erase_source}
    if layer:
        params["layer"] = layer
    return await _call("create_region", params)


@mcp.tool()
async def create_boundary(point: list[float], detect_islands: bool = True, layer: str = "") -> str:
    """Trace a closed boundary polyline around a point (AutoCAD BPOLY).

    The point must sit inside a fully enclosed area.
    """
    params: dict = {"point": point, "detect_islands": detect_islands}
    if layer:
        params["layer"] = layer
    return await _call("create_boundary", params)


@mcp.tool()
async def fillet_entities(id1: str, id2: str, radius: float) -> str:
    """Fillet two lines with a tangent arc, trimming both back to the tangent points.

    Currently supports two Line entities; use execute_command('FILLET') for other types.
    """
    return await _call("fillet_entities", {"id1": id1, "id2": id2, "radius": radius})


@mcp.tool()
async def overkill(tolerance: float = 1e-6, ignore_layer: bool = False,
                   __confirm: bool = False) -> str:
    """Delete exact duplicate/overlapping entities in model space (AutoCAD OVERKILL).

    Compares lines, circles, arcs, points and polylines. Destructive — pass
    __confirm=true to proceed.
    """
    return await _call("overkill", {
        "tolerance": tolerance, "ignore_layer": ignore_layer, "__confirm": __confirm,
    })


# =============================================================================
# Additional 2D Entities
# =============================================================================

@mcp.tool()
async def create_point(position: list[float], layer: str = "", color: int = -1) -> str:
    """Create a point entity."""
    params: dict = {"position": position}
    if layer:
        params["layer"] = layer
    if color >= 0:
        params["color"] = color
    return await _call("create_point", params)


@mcp.tool()
async def create_xline(point: list[float], through: list[float] | None = None,
                       angle: float = 0.0, layer: str = "") -> str:
    """Create an infinite construction line through a point, by angle or a second point."""
    params: dict = {"point": point, "angle": angle}
    if through:
        params["through"] = through
    if layer:
        params["layer"] = layer
    return await _call("create_xline", params)


@mcp.tool()
async def create_ray(point: list[float], through: list[float] | None = None,
                     angle: float = 0.0, layer: str = "") -> str:
    """Create a ray (semi-infinite line) from a point."""
    params: dict = {"point": point, "angle": angle}
    if through:
        params["through"] = through
    if layer:
        params["layer"] = layer
    return await _call("create_ray", params)


@mcp.tool()
async def create_polygon(center: list[float], sides: int, radius: float,
                         mode: str = "inscribed", rotation: float = 0.0,
                         layer: str = "", color: int = -1) -> str:
    """Create a regular polygon. mode: inscribed (radius to vertex) or circumscribed."""
    params: dict = {"center": center, "sides": sides, "radius": radius,
                    "mode": mode, "rotation": rotation}
    if layer:
        params["layer"] = layer
    if color >= 0:
        params["color"] = color
    return await _call("create_polygon", params)


@mcp.tool()
async def create_donut(center: list[float], outer_diameter: float, inner_diameter: float,
                       layer: str = "", color: int = -1) -> str:
    """Create a donut (filled annulus). inner_diameter=0 gives a filled dot."""
    params: dict = {"center": center, "outer_diameter": outer_diameter,
                    "inner_diameter": inner_diameter}
    if layer:
        params["layer"] = layer
    if color >= 0:
        params["color"] = color
    return await _call("create_donut", params)


@mcp.tool()
async def create_3d_polyline(points: list[list[float]], closed: bool = False,
                             layer: str = "") -> str:
    """Create a 3D polyline through [x,y,z] points."""
    params: dict = {"points": points, "closed": closed}
    if layer:
        params["layer"] = layer
    return await _call("create_3d_polyline", params)


# =============================================================================
# 3D Solids
# =============================================================================

@mcp.tool()
async def create_box(center: list[float], length: float, width: float, height: float,
                     layer: str = "") -> str:
    """Create a 3D box solid centred on a point."""
    params: dict = {"center": center, "length": length, "width": width, "height": height}
    if layer:
        params["layer"] = layer
    return await _call("create_box", params)


@mcp.tool()
async def create_sphere(center: list[float], radius: float, layer: str = "") -> str:
    """Create a 3D sphere solid."""
    params: dict = {"center": center, "radius": radius}
    if layer:
        params["layer"] = layer
    return await _call("create_sphere", params)


@mcp.tool()
async def create_cylinder(center: list[float], radius: float, height: float,
                          layer: str = "") -> str:
    """Create a 3D cylinder solid centred on a point."""
    params: dict = {"center": center, "radius": radius, "height": height}
    if layer:
        params["layer"] = layer
    return await _call("create_cylinder", params)


@mcp.tool()
async def create_cone(center: list[float], radius: float, height: float,
                      top_radius: float = 0.0, layer: str = "") -> str:
    """Create a 3D cone; top_radius > 0 makes a truncated cone (frustum)."""
    params: dict = {"center": center, "radius": radius, "height": height,
                    "top_radius": top_radius}
    if layer:
        params["layer"] = layer
    return await _call("create_cone", params)


@mcp.tool()
async def create_wedge(center: list[float], length: float, width: float, height: float,
                       layer: str = "") -> str:
    """Create a 3D wedge solid."""
    params: dict = {"center": center, "length": length, "width": width, "height": height}
    if layer:
        params["layer"] = layer
    return await _call("create_wedge", params)


@mcp.tool()
async def create_torus(center: list[float], major_radius: float, minor_radius: float,
                       layer: str = "") -> str:
    """Create a 3D torus solid."""
    params: dict = {"center": center, "major_radius": major_radius,
                    "minor_radius": minor_radius}
    if layer:
        params["layer"] = layer
    return await _call("create_torus", params)


@mcp.tool()
async def extrude_profile(ids: list[str], height: float, taper_angle: float = 0.0,
                          erase_source: bool = True) -> str:
    """Extrude closed profile curves into a 3D solid."""
    return await _call("extrude_profile", {
        "ids": ids, "height": height, "taper_angle": taper_angle,
        "erase_source": erase_source,
    })


@mcp.tool()
async def revolve_profile(ids: list[str], axis_point: list[float],
                          axis_direction: list[float] | None = None,
                          angle: float = 360.0, erase_source: bool = True) -> str:
    """Revolve closed profile curves around an axis into a 3D solid."""
    params: dict = {"ids": ids, "axis_point": axis_point, "angle": angle,
                    "erase_source": erase_source}
    if axis_direction:
        params["axis_direction"] = axis_direction
    return await _call("revolve_profile", params)


@mcp.tool()
async def boolean_solids(target: str, others: list[str], operation: str = "union") -> str:
    """Combine 3D solids. operation: union, subtract or intersect.

    The result is written into target; the other solids are consumed.
    """
    return await _call("boolean_solids",
                       {"target": target, "others": others, "operation": operation})


@mcp.tool()
async def get_solid_properties(id: str) -> str:
    """Get a 3D solid's volume, centroid, moments of inertia and bounding box."""
    return await _call("get_solid_properties", {"id": id})


# =============================================================================
# Groups, Layer States, Views, UCS
# =============================================================================

@mcp.tool()
async def create_group(name: str, ids: list[str], description: str = "") -> str:
    """Create a named group from entities so they select together."""
    params: dict = {"name": name, "ids": ids}
    if description:
        params["description"] = description
    return await _call("create_group", params)


@mcp.tool()
async def list_groups() -> str:
    """List named groups and their member counts."""
    return await _call("list_groups")


@mcp.tool()
async def add_to_group(name: str, ids: list[str], remove: bool = False) -> str:
    """Add entities to a group, or remove them with remove=true."""
    return await _call("add_to_group", {"name": name, "ids": ids, "remove": remove})


@mcp.tool()
async def ungroup(name: str, __confirm: bool = False) -> str:
    """Remove a group definition; its entities stay in the drawing.
    Destructive — pass __confirm=true to proceed."""
    return await _call("ungroup", {"name": name, "__confirm": __confirm})


@mcp.tool()
async def save_layer_state(name: str, description: str = "", overwrite: bool = False) -> str:
    """Save current layer visibility/colour/lock settings as a named layer state."""
    params: dict = {"name": name, "overwrite": overwrite}
    if description:
        params["description"] = description
    return await _call("save_layer_state", params)


@mcp.tool()
async def restore_layer_state(name: str) -> str:
    """Restore a previously saved layer state."""
    return await _call("restore_layer_state", {"name": name})


@mcp.tool()
async def list_layer_states() -> str:
    """List saved layer states."""
    return await _call("list_layer_states")


@mcp.tool()
async def delete_layer_state(name: str, __confirm: bool = False) -> str:
    """Delete a saved layer state. Destructive — pass __confirm=true to proceed."""
    return await _call("delete_layer_state", {"name": name, "__confirm": __confirm})


@mcp.tool()
async def create_named_view(name: str, min: list[float], max: list[float]) -> str:
    """Save a named view from a rectangular window (min and max corners)."""
    return await _call("create_named_view", {"name": name, "min": min, "max": max})


@mcp.tool()
async def list_named_views() -> str:
    """List saved named views."""
    return await _call("list_named_views")


@mcp.tool()
async def restore_view(name: str) -> str:
    """Restore a saved named view in the active viewport."""
    return await _call("restore_view", {"name": name})


@mcp.tool()
async def list_ucs() -> str:
    """List named user coordinate systems."""
    return await _call("list_ucs")


@mcp.tool()
async def set_ucs(name: str = "", origin: list[float] | None = None,
                  x_axis: list[float] | None = None, y_axis: list[float] | None = None,
                  save_as: str = "") -> str:
    """Set the active UCS.

    Pass name to activate a saved UCS ('World' resets), or origin (+ optional
    x_axis/y_axis) to define one, optionally saving it as save_as.
    """
    params: dict = {}
    if name:
        params["name"] = name
    if origin:
        params["origin"] = origin
    if x_axis:
        params["x_axis"] = x_axis
    if y_axis:
        params["y_axis"] = y_axis
    if save_as:
        params["save_as"] = save_as
    return await _call("set_ucs", params)


# =============================================================================
# Drawing Data & Audit
# =============================================================================

@mcp.tool()
async def get_xdata(id: str, app_name: str = "") -> str:
    """Read extended entity data (XData) attached to an entity."""
    params: dict = {"id": id}
    if app_name:
        params["app_name"] = app_name
    return await _call("get_xdata", params)


@mcp.tool()
async def set_xdata(id: str, app_name: str, values: list) -> str:
    """Attach extended entity data (XData) to an entity under an application name.

    Registers the application name automatically. Useful for tagging entities
    with your own metadata that survives in the DWG.
    """
    return await _call("set_xdata", {"id": id, "app_name": app_name, "values": values})


@mcp.tool()
async def get_drawing_properties() -> str:
    """Read drawing properties (title, author, subject, keywords, custom fields)."""
    return await _call("get_drawing_properties")


@mcp.tool()
async def set_drawing_properties(
    title: str = "",
    subject: str = "",
    author: str = "",
    keywords: str = "",
    comments: str = "",
    revision_number: str = "",
    hyperlink_base: str = "",
    custom: dict | None = None,
) -> str:
    """Set drawing properties, including custom name/value fields used by title blocks."""
    params: dict = {}
    for k, v in (
        ("title", title), ("subject", subject), ("author", author),
        ("keywords", keywords), ("comments", comments),
        ("revision_number", revision_number), ("hyperlink_base", hyperlink_base),
    ):
        if v:
            params[k] = v
    if custom:
        params["custom"] = custom
    return await _call("set_drawing_properties", params)


@mcp.tool()
async def entity_count_report(by_layer: bool = True, space: str = "model") -> str:
    """Count entities by type and by layer. space: model, paper or current."""
    return await _call("entity_count_report", {"by_layer": by_layer, "space": space})


@mcp.tool()
async def audit_drawing() -> str:
    """Health-check the drawing: empty layers, unused blocks, broken xrefs,
    zero-length curves, frozen/locked layers."""
    return await _call("audit_drawing")


# =============================================================================
# Multileaders
# =============================================================================

@mcp.tool()
async def create_multileader(
    arrow_point: list[float],
    text_point: list[float],
    text: str,
    height: float = 0.0,
    style: str = "",
    text_style: str = "",
    layer: str = "",
) -> str:
    """Create a multileader: an arrow at arrow_point with text at text_point."""
    params: dict = {"arrow_point": arrow_point, "text_point": text_point, "text": text}
    if height > 0:
        params["height"] = height
    if style:
        params["style"] = style
    if text_style:
        params["text_style"] = text_style
    if layer:
        params["layer"] = layer
    return await _call("create_multileader", params)


@mcp.tool()
async def list_mleader_styles() -> str:
    """List multileader styles with text height, arrow size and landing gap."""
    return await _call("list_mleader_styles")


@mcp.tool()
async def create_mleader_style(
    name: str,
    text_height: float = 0.0,
    arrow_size: float = 0.0,
    landing_gap: float = -1.0,
    text_style: str = "",
) -> str:
    """Create a multileader style."""
    params: dict = {"name": name}
    if text_height > 0:
        params["text_height"] = text_height
    if arrow_size > 0:
        params["arrow_size"] = arrow_size
    if landing_gap >= 0:
        params["landing_gap"] = landing_gap
    if text_style:
        params["text_style"] = text_style
    return await _call("create_mleader_style", params)


# =============================================================================
# Additional Dimension Types
# =============================================================================

@mcp.tool()
async def create_ordinate_dimension(
    point: list[float],
    leader_end: list[float],
    axis: str = "x",
    text: str = "",
    style: str = "",
    layer: str = "",
) -> str:
    """Create an ordinate dimension measuring along the x or y axis from the UCS origin."""
    params: dict = {"point": point, "leader_end": leader_end, "axis": axis}
    if text:
        params["text"] = text
    if style:
        params["style"] = style
    if layer:
        params["layer"] = layer
    return await _call("create_ordinate_dimension", params)


@mcp.tool()
async def create_arclength_dimension(
    center: list[float],
    start: list[float],
    end: list[float],
    arc_point: list[float],
    text: str = "",
    style: str = "",
    layer: str = "",
) -> str:
    """Create an arc-length dimension for an arc defined by center, start and end."""
    params: dict = {"center": center, "start": start, "end": end, "arc_point": arc_point}
    if text:
        params["text"] = text
    if style:
        params["style"] = style
    if layer:
        params["layer"] = layer
    return await _call("create_arclength_dimension", params)


@mcp.tool()
async def create_tolerance(
    text: str,
    position: list[float],
    height: float = 0.0,
    layer: str = "",
) -> str:
    """Create a GD&T feature control frame (tolerance symbol).

    text uses AutoCAD's tolerance encoding, with %%v separating compartments,
    e.g. "{\\Fgdt;j}%%v{\\Fgdt;n}0.05%%v%%v%%v%%v%%v".
    """
    params: dict = {"text": text, "position": position}
    if height > 0:
        params["height"] = height
    if layer:
        params["layer"] = layer
    return await _call("create_tolerance", params)


@mcp.tool()
async def edit_dimension_text(id: str, text: str) -> str:
    """Override a dimension's displayed text.

    Pass "" to clear the override and restore the measured value, or include
    <> to embed the measurement (e.g. "<> TYP").
    """
    return await _call("edit_dimension_text", {"id": id, "text": text})


@mcp.tool()
async def update_dimensions(ids: list[str] | None = None, style: str = "") -> str:
    """Regenerate dimensions from current settings, optionally reassigning a style.

    Omit ids to update every dimension in the current space.
    """
    params: dict = {}
    if ids:
        params["ids"] = ids
    if style:
        params["style"] = style
    return await _call("update_dimensions", params)


# =============================================================================
# Annotation Scaling
# =============================================================================

@mcp.tool()
async def list_annotation_scales() -> str:
    """List annotation scales available in the drawing and the current one."""
    return await _call("list_annotation_scales")


@mcp.tool()
async def set_annotation_scale(name: str) -> str:
    """Set the current annotation scale (CANNOSCALE), e.g. "1:100"."""
    return await _call("set_annotation_scale", {"name": name})


@mcp.tool()
async def add_annotation_scale_to_entity(
    ids: list[str], scale: str, remove: bool = False
) -> str:
    """Add (or remove) an annotation scale representation on annotative entities.

    Makes entities annotative if they are not already. Entities that cannot be
    made annotative are reported in 'skipped' rather than failing the call.
    """
    return await _call("add_annotation_scale_to_entity",
                       {"ids": ids, "scale": scale, "remove": remove})


# =============================================================================
# Tables
# =============================================================================

@mcp.tool()
async def get_table_data(id: str) -> str:
    """Read a table's full contents as a 2D array of cell strings."""
    return await _call("get_table_data", {"id": id})


@mcp.tool()
async def set_table_cell(
    id: str,
    row: int | None = None,
    column: int | None = None,
    text: str = "",
    cells: list[dict] | None = None,
    text_height: float = 0.0,
) -> str:
    """Write table cells.

    Either row+column+text for a single cell, or cells as a list of
    {"row": r, "column": c, "text": "..."} to update many at once.
    """
    params: dict = {"id": id}
    if cells is not None:
        params["cells"] = cells
    else:
        params["row"] = row
        params["column"] = column
        params["text"] = text
        if text_height > 0:
            params["text_height"] = text_height
    return await _call("set_table_cell", params)


@mcp.tool()
async def merge_table_cells(
    id: str,
    top_row: int,
    bottom_row: int,
    left_column: int,
    right_column: int,
    unmerge: bool = False,
) -> str:
    """Merge (or unmerge) a rectangular range of table cells."""
    return await _call("merge_table_cells", {
        "id": id, "top_row": top_row, "bottom_row": bottom_row,
        "left_column": left_column, "right_column": right_column, "unmerge": unmerge,
    })


@mcp.tool()
async def list_table_styles() -> str:
    """List table styles defined in the drawing."""
    return await _call("list_table_styles")


# =============================================================================
# Text Editing, Wipeout, Revision Cloud
# =============================================================================

@mcp.tool()
async def edit_mtext(
    id: str,
    text: str = "",
    height: float = 0.0,
    width: float = 0.0,
    rotation: float | None = None,
    text_style: str = "",
) -> str:
    """Edit existing text or mtext in place — contents, height, width, rotation, style."""
    params: dict = {"id": id}
    if text:
        params["text"] = text
    if height > 0:
        params["height"] = height
    if width > 0:
        params["width"] = width
    if rotation is not None:
        params["rotation"] = rotation
    if text_style:
        params["text_style"] = text_style
    return await _call("edit_mtext", params)


@mcp.tool()
async def create_wipeout(points: list[list[float]], layer: str = "") -> str:
    """Create a wipeout mask from a closed polygon of at least 3 points.

    Frame visibility is global — control it with set_system_variable("WIPEOUTFRAME", ...).
    """
    params: dict = {"points": points}
    if layer:
        params["layer"] = layer
    return await _call("create_wipeout", params)


@mcp.tool()
async def create_revision_cloud(
    min: list[float],
    max: list[float],
    arc_length: float = 0.0,
    layer: str = "",
    color: int = -1,
) -> str:
    """Create a rectangular revision cloud between two opposite corners.

    arc_length controls bump size; omit it for a sensible default based on the
    rectangle size.
    """
    params: dict = {"min": min, "max": max}
    if arc_length > 0:
        params["arc_length"] = arc_length
    if layer:
        params["layer"] = layer
    if color >= 0:
        params["color"] = color
    return await _call("create_revision_cloud", params)


# =============================================================================
# Sheet Sets (COM-based, read-only)
# =============================================================================

@mcp.tool()
async def get_sheet_set_status() -> str:
    """Check whether Sheet Set Manager automation is available on this machine.

    Sheet sets have no managed .NET API, so these tools go through the
    AcSmComponents COM library. Call this before relying on the others.
    """
    return await _call("get_sheet_set_status")


@mcp.tool()
async def open_sheet_set(path: str) -> str:
    """Open a .dst sheet set file and report its name, description and sheet count."""
    return await _call("open_sheet_set", {"path": path})


@mcp.tool()
async def list_sheets(path: str) -> str:
    """List the sheets in a .dst sheet set with number, title, name and description."""
    return await _call("list_sheets", {"path": path})


@mcp.tool()
async def close_sheet_set(path: str) -> str:
    """Close an open sheet set database."""
    return await _call("close_sheet_set", {"path": path})


# =============================================================================
# Entry point
# =============================================================================

if __name__ == "__main__":
    mcp.run(transport="stdio")
